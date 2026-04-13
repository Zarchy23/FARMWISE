# core/export_service.py
"""
Unified Export Center Service
Export data in multiple formats (CSV, PDF, JSON, Excel)
"""

import csv
import json
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from typing import Dict, List, Any, Tuple
from decimal import Decimal
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors

from django.utils import timezone
from django.template.loader import render_to_string
from django.http import HttpResponse

from .models import Farm, CropSeason, Animal, EquipmentBooking, Transaction, ProductListing, Order, Supermarket


class ExportService:
    """Service for exporting data in various formats"""
    
    EXPORT_FORMATS = ['csv', 'json', 'pdf', 'xlsx', 'html']
    
    DATA_TYPES = [
        ('crops', 'Crop Data'),
        ('livestock', 'Livestock Data'),
        ('equipment', 'Equipment Rentals'),
        ('financial', 'Financial Data'),
        ('labor', 'Labor Records'),
        ('pest_reports', 'Pest Reports'),
        ('insurance', 'Insurance Policies'),
        ('all', 'All Data'),
    ]
    
    @staticmethod
    def export_crops(
        farm: Farm,
        start_date: datetime = None,
        end_date: datetime = None,
        format: str = 'csv'
    ) -> Tuple[HttpResponse, str]:
        """Export crop data"""
        
        # Build query with or without date filters
        crops_query = CropSeason.objects.filter(field__farm=farm).select_related('crop_type', 'field')
        
        if start_date is not None and end_date is not None:
            crops_query = crops_query.filter(
                planting_date__gte=start_date,
                planting_date__lte=end_date
            )
        
        crops = crops_query.all()
        
        data = []
        for crop in crops:
            data.append({
                'field_name': crop.field.name,
                'crop_type': crop.crop_type.name,
                'variety': crop.variety,
                'season': crop.get_season_display(),
                'planting_date': crop.planting_date.isoformat(),
                'harvest_date': crop.expected_harvest_date.isoformat(),
                'actual_harvest_date': crop.actual_harvest_date.isoformat() if crop.actual_harvest_date else '',
                'status': crop.get_status_display(),
                'estimated_yield_kg': float(crop.estimated_yield_kg or 0),
                'actual_yield_kg': float(crop.actual_yield_kg or 0),
                'estimated_revenue': float(crop.estimated_revenue or 0),
                'actual_revenue': float(crop.actual_revenue or 0),
            })
        
        return ExportService._format_export(data, format, 'crop_data')
    
    @staticmethod
    def export_livestock(
        farm: Farm,
        start_date: datetime = None,
        end_date: datetime = None,
        format: str = 'csv'
    ) -> Tuple[HttpResponse, str]:
        """Export livestock data"""
        
        if start_date is None:
            start_date = timezone.now() - timedelta(days=365)
        if end_date is None:
            end_date = timezone.now()
        
        animals = Animal.objects.filter(
            farm=farm,
            created_at__gte=start_date,
            created_at__lte=end_date
        ).select_related('animal_type')
        
        data = []
        for animal in animals:
            data.append({
                'tag_number': animal.tag_number,
                'name': animal.name,
                'species': animal.animal_type.get_species_display(),
                'breed': animal.animal_type.breed,
                'gender': animal.get_gender_display(),
                'birth_date': animal.birth_date.isoformat() if animal.birth_date else '',
                'purchase_date': animal.purchase_date.isoformat() if animal.purchase_date else '',
                'purchase_price': float(animal.purchase_price or 0),
                'weight_kg': float(animal.weight_kg or 0),
                'status': animal.get_status_display(),
                'location': animal.location,
            })
        
        return ExportService._format_export(data, format, 'livestock_data')
    
    @staticmethod
    def export_financial(
        farm: Farm,
        start_date: datetime = None,
        end_date: datetime = None,
        format: str = 'csv'
    ) -> Tuple[HttpResponse, str]:
        """Export financial data"""
        
        if start_date is None:
            start_date = timezone.now() - timedelta(days=365)
        if end_date is None:
            end_date = timezone.now()
        
        transactions = Transaction.objects.filter(
            farm=farm,
            created_at__gte=start_date,
            created_at__lte=end_date
        ).order_by('-created_at')
        
        data = []
        total_income = Decimal('0.00')
        total_expense = Decimal('0.00')
        
        for transaction in transactions:
            amount = float(transaction.amount)
            
            if transaction.transaction_type == 'income':
                total_income += transaction.amount
            else:
                total_expense += transaction.amount
            
            data.append({
                'date': transaction.created_at.date().isoformat(),
                'type': transaction.get_transaction_type_display(),
                'category': transaction.get_category_display(),
                'description': transaction.description,
                'amount': amount,
                'notes': transaction.notes,
            })
        
        # Add summary
        data.append({
            'date': '',
            'type': 'SUMMARY',
            'category': '',
            'description': 'Total Income',
            'amount': float(total_income),
            'notes': ''
        })
        data.append({
            'date': '',
            'type': 'SUMMARY',
            'category': '',
            'description': 'Total Expense',
            'amount': float(total_expense),
            'notes': ''
        })
        data.append({
            'date': '',
            'type': 'SUMMARY',
            'category': '',
            'description': 'Net Profit',
            'amount': float(total_income - total_expense),
            'notes': ''
        })
        
        return ExportService._format_export(data, format, 'financial_data')
    
    @staticmethod
    def export_equipment(user, format: str = 'csv') -> Tuple[HttpResponse, str]:
        """Export equipment data"""
        from .models import Equipment
        
        equipment = Equipment.objects.filter(owner=user).all()
        
        data = []
        for equip in equipment:
            data.append({
                'name': equip.name,
                'type': equip.equipment_type if hasattr(equip, 'equipment_type') else 'N/A',
                'brand': equip.brand if hasattr(equip, 'brand') else 'N/A',
                'model': equip.model if hasattr(equip, 'model') else 'N/A',
                'purchase_date': equip.purchase_date.isoformat() if hasattr(equip, 'purchase_date') and equip.purchase_date else 'N/A',
                'purchase_price': float(equip.purchase_price) if hasattr(equip, 'purchase_price') and equip.purchase_price else 0,
                'status': equip.status if hasattr(equip, 'status') else 'Active',
                'location': equip.location if hasattr(equip, 'location') else 'N/A',
            })
        
        return ExportService._format_export(data, format, 'equipment_data')
    
    @staticmethod
    def export_insurance(user, format: str = 'csv') -> Tuple[HttpResponse, str]:
        """Export insurance policies"""
        from .models import InsurancePolicy
        
        policies = InsurancePolicy.objects.filter(farm__owner=user).all()
        
        data = []
        for policy in policies:
            data.append({
                'farm': policy.farm.name if hasattr(policy, 'farm') and policy.farm else 'N/A',
                'policy_number': policy.policy_number if hasattr(policy, 'policy_number') else 'N/A',
                'provider': policy.provider if hasattr(policy, 'provider') else 'N/A',
                'coverage_type': policy.coverage_type if hasattr(policy, 'coverage_type') else 'N/A',
                'start_date': policy.start_date.isoformat() if hasattr(policy, 'start_date') and policy.start_date else 'N/A',
                'end_date': policy.end_date.isoformat() if hasattr(policy, 'end_date') and policy.end_date else 'N/A',
                'premium_amount': float(policy.premium_amount) if hasattr(policy, 'premium_amount') and policy.premium_amount else 0,
                'status': policy.status if hasattr(policy, 'status') else 'Active',
            })
        
        return ExportService._format_export(data, format, 'insurance_data')
    
    @staticmethod
    def export_payroll(user, format: str = 'csv') -> Tuple[HttpResponse, str]:
        """Export payroll records"""
        from .models import PayrollRecord
        
        payroll = PayrollRecord.objects.filter(farm__owner=user).all()
        
        data = []
        for record in payroll:
            data.append({
                'farm': record.farm.name if hasattr(record, 'farm') and record.farm else 'N/A',
                'employee': str(record.employee) if hasattr(record, 'employee') and record.employee else 'N/A',
                'pay_period': str(record.pay_period) if hasattr(record, 'pay_period') else 'N/A',
                'base_salary': float(record.base_salary) if hasattr(record, 'base_salary') and record.base_salary else 0,
                'deductions': float(record.deductions) if hasattr(record, 'deductions') and record.deductions else 0,
                'net_salary': float(record.net_salary) if hasattr(record, 'net_salary') and record.net_salary else 0,
                'payment_date': record.payment_date.isoformat() if hasattr(record, 'payment_date') and record.payment_date else 'N/A',
                'status': record.status if hasattr(record, 'status') else 'Paid',
            })
        
        return ExportService._format_export(data, format, 'payroll_data')
    
    @staticmethod
    def export_pest_reports(user, format: str = 'csv') -> Tuple[HttpResponse, str]:
        """Export pest detection reports"""
        from .models import PestDetectionReport
        
        reports = PestDetectionReport.objects.filter(farm__owner=user).all()
        
        data = []
        for report in reports:
            data.append({
                'farm': report.farm.name if hasattr(report, 'farm') and report.farm else 'N/A',
                'detection_date': report.detection_date.isoformat() if hasattr(report, 'detection_date') and report.detection_date else 'N/A',
                'pest_name': report.pest_name if hasattr(report, 'pest_name') else 'N/A',
                'severity': report.severity if hasattr(report, 'severity') else 'Low',
                'area_affected': str(report.area_affected) if hasattr(report, 'area_affected') else '0',
                'treatment': report.treatment if hasattr(report, 'treatment') else 'N/A',
                'status': report.status if hasattr(report, 'status') else 'Reported',
            })
        
        return ExportService._format_export(data, format, 'pest_reports_data')
    
    @staticmethod
    def export_carbon(user, format: str = 'csv') -> Tuple[HttpResponse, str]:
        """Export carbon footprint data"""
        from .models import CarbonFootprint
        
        carbon_data = CarbonFootprint.objects.filter(farm__owner=user).all()
        
        data = []
        for record in carbon_data:
            data.append({
                'farm': record.farm.name if hasattr(record, 'farm') and record.farm else 'N/A',
                'period': str(record.period) if hasattr(record, 'period') else 'N/A',
                'source': record.source if hasattr(record, 'source') else 'N/A',
                'emissions_kg_co2': float(record.emissions_kg_co2) if hasattr(record, 'emissions_kg_co2') and record.emissions_kg_co2 else 0,
                'mitigation_action': record.mitigation_action if hasattr(record, 'mitigation_action') else 'N/A',
                'reduction_target': float(record.reduction_target) if hasattr(record, 'reduction_target') and record.reduction_target else 0,
            })
        
        return ExportService._format_export(data, format, 'carbon_data')
    
    @staticmethod
    def export_reminders(user, format: str = 'csv') -> Tuple[HttpResponse, str]:
        """Export reminders"""
        from .models import Reminder
        
        reminders = Reminder.objects.filter(farm__owner=user).all()
        
        data = []
        for remind in reminders:
            data.append({
                'farm': remind.farm.name if hasattr(remind, 'farm') and remind.farm else 'N/A',
                'reminder_type': remind.reminder_type if hasattr(remind, 'reminder_type') else 'N/A',
                'title': remind.title if hasattr(remind, 'title') else 'N/A',
                'description': remind.description if hasattr(remind, 'description') else 'N/A',
                'due_date': remind.due_date.isoformat() if hasattr(remind, 'due_date') and remind.due_date else 'N/A',
                'status': 'Completed' if remind.is_completed else 'Pending',
            })
        
        return ExportService._format_export(data, format, 'reminders_data')
    
    @staticmethod
    def _format_export(
        data: List[Dict[str, Any]],
        format: str,
        filename_base: str
    ) -> Tuple[HttpResponse, str]:
        """Format data for export in specified format"""
        
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'csv':
            return ExportService._export_csv(data, f'{filename_base}_{timestamp}.csv')
        
        elif format == 'json':
            return ExportService._export_json(data, f'{filename_base}_{timestamp}.json')
        
        elif format == 'xlsx':
            return ExportService._export_xlsx(data, f'{filename_base}_{timestamp}.xlsx')
        
        elif format == 'pdf':
            return ExportService._export_pdf(data, f'{filename_base}_{timestamp}.pdf')
        
        elif format == 'html':
            return ExportService._export_html(data, f'{filename_base}_{timestamp}.html')
        
        else:
            raise ValueError(f'Unsupported format: {format}')
    
    @staticmethod
    def _export_csv(data: List[Dict], filename: str) -> Tuple[HttpResponse, str]:
        """Export to CSV"""
        
        if not data:
            data = [{}]
        
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response, filename
    
    @staticmethod
    def _export_json(data: List[Dict], filename: str) -> Tuple[HttpResponse, str]:
        """Export to JSON"""
        
        json_data = json.dumps(data, indent=2, default=str)
        
        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response, filename
    
    @staticmethod
    def _export_xlsx(data: List[Dict], filename: str) -> Tuple[HttpResponse, str]:
        """Export to Excel"""
        
        if not data:
            data = [{}]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data'
        
        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response, filename
    
    @staticmethod
    def _export_pdf(data: List[Dict], filename: str) -> Tuple[HttpResponse, str]:
        """Export to PDF"""
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title = Paragraph(f"Farm Report - {timezone.now().date()}", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.5*inch))
        
        # Create table
        if data:
            headers = list(data[0].keys())
            table_data = [headers] + [[str(row.get(h, '')) for h in headers] for row in data]
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response, filename
    
    @staticmethod
    def _export_html(data: List[Dict], filename: str) -> Tuple[HttpResponse, str]:
        """Export to HTML"""
        
        html_content = '<table border="1" cellpadding="5"><tr>'
        
        if data:
            # Headers
            for header in data[0].keys():
                html_content += f'<th>{header}</th>'
            html_content += '</tr>'
            
            # Rows
            for row in data:
                html_content += '<tr>'
                for header in data[0].keys():
                    html_content += f'<td>{row.get(header, "")}</td>'
                html_content += '</tr>'
        
        html_content += '</table>'
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response, filename
    
    # ============================================================
    # MARKETPLACE EXPORT METHODS - FOR SUPERMARKETS
    # ============================================================
    
    @staticmethod
    def export_marketplace_products(
        user,
        start_date: datetime = None,
        end_date: datetime = None,
        format: str = 'csv',
        status_filter: str = None
    ) -> Tuple[HttpResponse, str]:
        """Export marketplace products for supermarket/seller"""
        
        if start_date is None:
            start_date = timezone.now() - timedelta(days=365)
        if end_date is None:
            end_date = timezone.now()
        
        # Get products based on user type
        if user.user_type == 'supermarket':
            try:
                supermarket = user.supermarket_profile
                products = ProductListing.objects.filter(
                    seller__owner=user,
                    created_at__gte=start_date,
                    created_at__lte=end_date
                )
            except Supermarket.DoesNotExist:
                products = ProductListing.objects.none()
        else:
            # Farmer or other sellers
            products = ProductListing.objects.filter(
                seller__owner=user,
                created_at__gte=start_date,
                created_at__lte=end_date
            )
        
        # Apply status filter if provided
        if status_filter and status_filter != 'all':
            products = products.filter(status=status_filter)
        
        data = []
        for product in products:
            data.append({
                'product_name': product.product_name,
                'category': product.category,
                'quantity': float(product.quantity),
                'unit': product.unit,
                'price_per_unit': float(product.price_per_unit),
                'total_price': float(product.total_price),
                'status': product.get_status_display(),
                'is_organic': 'Yes' if product.is_organic else 'No',
                'is_out_of_stock': 'Yes' if product.is_out_of_stock else 'No',
                'delivery_available': 'Yes' if product.delivery_available else 'No',
                'delivery_fee': float(product.delivery_fee),
                'created_at': product.created_at.date().isoformat(),
                'harvest_date': product.harvest_date.isoformat() if product.harvest_date else '',
                'expiry_date': product.expiry_date.isoformat() if product.expiry_date else '',
            })
        
        return ExportService._format_export(data, format, 'marketplace_products')
    
    @staticmethod
    def export_marketplace_sales(
        user,
        start_date: datetime = None,
        end_date: datetime = None,
        format: str = 'csv'
    ) -> Tuple[HttpResponse, str]:
        """Export sales/orders data for marketplace seller"""
        
        if start_date is None:
            start_date = timezone.now() - timedelta(days=365)
        if end_date is None:
            end_date = timezone.now()
        
        # Get orders where user is the seller
        orders = Order.objects.filter(
            listing__seller__owner=user,
            created_at__gte=start_date,
            created_at__lte=end_date
        ).select_related('listing', 'buyer', 'listing__seller')
        
        data = []
        total_revenue = Decimal('0.00')
        
        for order in orders:
            amount = float(order.quantity * order.price_per_unit)
            total_revenue += order.quantity * order.price_per_unit
            
            data.append({
                'order_id': order.id,
                'buyer_name': order.buyer.get_full_name() or order.buyer.username,
                'buyer_phone': order.buyer.phone_number,
                'product_name': order.listing.product_name,
                'quantity': float(order.quantity),
                'unit': order.listing.unit,
                'price_per_unit': float(order.price_per_unit),
                'total_amount': amount,
                'status': order.get_status_display(),
                'delivery_required': 'Yes' if order.delivery_required else 'No',
                'delivery_fee': float(order.delivery_fee or 0),
                'order_date': order.created_at.date().isoformat(),
                'delivery_date': order.desired_delivery_date.isoformat() if order.desired_delivery_date else '',
                'notes': order.notes,
            })
        
        # Add summary
        data.append({
            'order_id': '',
            'buyer_name': 'SUMMARY',
            'buyer_phone': '',
            'product_name': 'Total Revenue',
            'quantity': '',
            'unit': '',
            'price_per_unit': '',
            'total_amount': float(total_revenue),
            'status': '',
            'delivery_required': '',
            'delivery_fee': '',
            'order_date': '',
            'delivery_date': '',
            'notes': '',
        })
        
        return ExportService._format_export(data, format, 'marketplace_sales')
    
    @staticmethod
    def export_marketplace_inventory_report(
        user,
        format: str = 'csv'
    ) -> Tuple[HttpResponse, str]:
        """Export inventory report for supermarket"""
        
        # Get all active products
        products = ProductListing.objects.filter(
            seller__owner=user,
            status='active'
        ).select_related('seller')
        
        data = []
        total_inventory_value = Decimal('0.00')
        out_of_stock_count = 0
        low_stock_warning = Decimal('10')  # Warning threshold
        
        for product in products:
            inventory_value = product.quantity * product.price_per_unit
            total_inventory_value += inventory_value
            
            stock_status = 'Out of Stock'
            if not product.is_out_of_stock:
                if product.quantity < low_stock_warning:
                    stock_status = 'Low Stock'
                else:
                    stock_status = 'In Stock'
            else:
                out_of_stock_count += 1
            
            data.append({
                'product_name': product.product_name,
                'category': product.category,
                'quantity_available': float(product.quantity),
                'unit': product.unit,
                'price_per_unit': float(product.price_per_unit),
                'inventory_value': float(inventory_value),
                'stock_status': stock_status,
                'is_organic': 'Yes' if product.is_organic else 'No',
                'harvest_date': product.harvest_date.isoformat() if product.harvest_date else 'N/A',
                'days_in_inventory': (timezone.now().date() - product.created_at.date()).days,
                'last_updated': product.updated_at.date().isoformat(),
            })
        
        # Add summary
        data.append({
            'product_name': 'SUMMARY',
            'category': '',
            'quantity_available': '',
            'unit': '',
            'price_per_unit': '',
            'inventory_value': float(total_inventory_value),
            'stock_status': f'Total Items: {products.count()}, Out of Stock: {out_of_stock_count}',
            'is_organic': '',
            'harvest_date': '',
            'days_in_inventory': '',
            'last_updated': timezone.now().date().isoformat(),
        })
        
        return ExportService._format_export(data, format, 'inventory_report')
    
    @staticmethod
    def export_supermarket_transactions(
        user,
        start_date: datetime = None,
        end_date: datetime = None,
        format: str = 'csv'
    ) -> Tuple[HttpResponse, str]:
        """Export financial transactions for supermarket"""
        
        if start_date is None:
            start_date = timezone.now() - timedelta(days=365)
        if end_date is None:
            end_date = timezone.now()
        
        # Get all farms owned by supermarket (if applicable)
        farms = Farm.objects.filter(owner=user)
        transactions = Transaction.objects.filter(
            farm__in=farms,
            created_at__gte=start_date,
            created_at__lte=end_date
        ).order_by('-created_at')
        
        data = []
        total_income = Decimal('0.00')
        total_expense = Decimal('0.00')
        
        for transaction in transactions:
            amount = float(transaction.amount)
            
            if transaction.transaction_type == 'income':
                total_income += transaction.amount
            else:
                total_expense += transaction.amount
            
            data.append({
                'date': transaction.created_at.date().isoformat(),
                'type': transaction.get_transaction_type_display(),
                'category': transaction.get_category_display(),
                'description': transaction.description,
                'amount': amount,
                'reference': transaction.reference,
                'farm': transaction.farm.name if transaction.farm else '',
            })
        
        # Add summary
        net_profit = total_income - total_expense
        data.append({
            'date': '',
            'type': 'SUMMARY',
            'category': '',
            'description': 'Total Income',
            'amount': float(total_income),
            'reference': '',
            'farm': '',
        })
        data.append({
            'date': '',
            'type': 'SUMMARY',
            'category': '',
            'description': 'Total Expense',
            'amount': float(total_expense),
            'reference': '',
            'farm': '',
        })
        data.append({
            'date': '',
            'type': 'SUMMARY',
            'category': '',
            'description': 'Net Profit/Loss',
            'amount': float(net_profit),
            'reference': '',
            'farm': '',
        })
        
        return ExportService._format_export(data, format, 'supermarket_transactions')
